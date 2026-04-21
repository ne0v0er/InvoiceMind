from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import Invoice

HEADERS = [
    "Société",
    "N° Facture",
    "Date Facture",
    "Date Échéance",
    "TVA 5.5%",
    "TVA 10%",
    "TVA 20%",
    "Total TTC",
]

CURRENCY_FMT = '#,##0.00 "€"'
HEADER_COLOR = "4472C4"
TOTAL_COLOR = "D9E1F2"
CURRENCY_COLUMNS = (5, 6, 7, 8)


def write_excel(invoices: list[Invoice], output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    _write_header(ws)

    for row_idx, invoice in enumerate(invoices, start=2):
        _write_invoice_row(ws, row_idx, invoice)

    data_last_row = len(invoices) + 1
    _write_total_row(ws, data_last_row + 1)

    _set_column_widths(ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"invoices_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


def _write_header(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)

    for col, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _write_invoice_row(ws, row: int, invoice: Invoice) -> None:
    ws.cell(row=row, column=1, value=invoice.company_name)
    ws.cell(row=row, column=2, value=invoice.invoice_number)
    ws.cell(row=row, column=3, value=invoice.invoice_date)
    ws.cell(row=row, column=4, value=invoice.due_date)
    ws.cell(row=row, column=5, value=invoice.tax_5_5)
    ws.cell(row=row, column=6, value=invoice.tax_10)
    ws.cell(row=row, column=7, value=invoice.tax_20)
    ws.cell(row=row, column=8, value=invoice.total_amount)

    for col in CURRENCY_COLUMNS:
        ws.cell(row=row, column=col).number_format = CURRENCY_FMT


def _write_total_row(ws, row: int) -> None:
    total_fill = PatternFill("solid", fgColor=TOTAL_COLOR)
    total_font = Font(bold=True)
    data_last_row = row - 1

    ws.cell(row=row, column=1, value="TOTAL").font = total_font

    for col in CURRENCY_COLUMNS:
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=row,
            column=col,
            value=f"=SUM({col_letter}2:{col_letter}{data_last_row})",
        )
        cell.number_format = CURRENCY_FMT
        cell.font = total_font
        cell.fill = total_fill


def _set_column_widths(ws) -> None:
    widths = [30, 18, 16, 16, 14, 14, 14, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
